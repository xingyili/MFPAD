"""
特征矩阵
1. HA读取->列表
2. 计算12中特征
"""
import pandas as pd
import re
from openpyxl import Workbook
import xlrd
import openpyxl
#1.读取HA序列文件，放入列表中
# def readHA():
# #     f = open("270_328acc.fasta")
#     f=open("F:/毕设/毕设-完整版/new/h3n2v.fasta")
#     line = f.readline()
#     while line:
#         line = f.readline()
#         HAs.append(line)
#         line = f.readline()
# #
def readHA():

    # HAs=[]
    # table = xlrd.open_workbook(filename)
    # sheet = table.sheet_by_index(0)
    # for i in range(sheet.nrows):
    #     HAs.append(sheet.cell(i,2).value)

    # table = xlrd.open_workbook("virus_sequence_270.xlsx")
    # sheet = table.sheet_by_index(0)
    # for i in range(sheet.nrows):
    #     HAs.append(sheet.cell(i, 1).value)
    #
    # table = xlrd.open_workbook("virus_2022_sequence.xlsx")
    # sheet = table.sheet_by_index(0)
    # for i in range(sheet.nrows):
    #     HAs.append(sheet.cell(i, 2).value)

    # table = xlrd.open_workbook("HA_2022_now.xls")
    # sheet = table.sheet_by_index(0)
    # for i in range(sheet.nrows):
    #     HAs_1.append(sheet.cell(i,2).value)
    # table = xlrd.open_workbook("vaccine_strain.xls")
    # sheet = table.sheet_by_index(0)
    # for i in range(sheet.nrows):
    #     HAs_2.append(sheet.cell(i, 2).value)



        # 读取 "HA_2022_now.xlsx"
    ha_2022_now_file = "HA_2022_now.xlsx"
    df_ha_2022_now = pd.read_excel(ha_2022_now_file)

    # 提取第一列数据
    HAs_1 = df_ha_2022_now.iloc[:, 1].tolist()

    # 读取 "vaccine_strain.xlsx"
    vaccine_strain_file = "vaccine_strain_final.xlsx"
    df_vaccine_strain = pd.read_excel(vaccine_strain_file)

    # 提取第一列数据
    HAs_2 = df_vaccine_strain.iloc[:, 1].tolist()

    return HAs_1,HAs_2


#2.计算12种特征
"""
两序列突变的数量
重要位点是否发生突变
5大抗原表位是否发生突变
糖基化位点是否发生突变
4种氨基酸理化性质
"""

#特征矩阵
def feature_cal(HAs_1,HAs_2):
    Feature=[]
    # for i in range(0,1):
    print(len(HAs_1))
    print(len(HAs_2))
    for i in range(0,len(HAs_1)):
        # print(i)
        for j in range(0,len(HAs_2)):
            fea=[]
            f=feature_0_1(HAs_1[i],HAs_2[j])
            for k in range(328):
                fea.append(f[k])
        # 2.1 序列突变数量
            mut_pos=mutation_number(HAs_1[i],HAs_2[j])
            f1_mutnum=len(mut_pos)
            fea.append(f1_mutnum)
        #2.2 重要位点是否发生突变
            f2_imsite=im_site(HAs_1[i],HAs_2[j])
            fea.append(f2_imsite)
        #2.3 五大抗原表位是否发生突变
            f3_A = epitopes(HAs_1[i][140:147],HAs_2[j][140:147])
            f4_B = epitopes(HAs_1[i][155:161]+HAs_1[i][187:197], HAs_2[j][155:161]+HAs_2[j][187:197])
            f5_C = epitopes(HAs_1[i][260:265], HAs_2[j][260:265])  #E
            f6_D = epitopes(HAs_1[i][275:281], HAs_2[j][275:281])  #C
            f7_E = epitopes(HAs_1[i][207:213], HAs_2[j][207:213])  #D
            fea.append(f3_A)
            fea.append(f4_B)
            fea.append(f5_C)
            fea.append(f6_D)
            fea.append(f7_E)
        #2.4 糖基化位点是否发生突变
            f8_gly = gly_site(HAs_1[i],HAs_2[j])
            fea.append(f8_gly)
        #2.5 4种理化性质
            f9_12=phy_pro(HAs_1[i],HAs_2[j],mut_pos)
            fea=fea+f9_12

            Feature.append(fea)
    return Feature

def feature_0_1(ha1,ha2):
    f=[]
    for i in range(328):
        if ha1[i]==ha2[i]:
            f.append(0)
        elif ha1[i]!=ha2[i]:
            f.append(1)
    return f



#序列突变数量
def mutation_number(ha1,ha2):
    mut_pos=[]
    for i in range(len(ha1)-1):
        if ha1[i] != ha2[i] and ha1[i]!="-" and ha2[i]!="-" and ha1[i]!="X" and ha2[i]!="X" :
            mut_pos.append(i)

    return mut_pos

#重要位点是否发生突变
def im_site(ha1,ha2):
    imsites=[128,130,131,132,142,144,145,155,156,157,158,159,160,189,193]
    is_mu=0
    for i in imsites:
        if ha1[i]!=ha2[i] and ha1[i]!="-" and ha2[i]!="-"and ha1[i]!="X" and ha2[i]!="X":
            is_mu=1
            break
    return is_mu

#抗原表位是否发生变化
def epitopes(ha1,ha2):
    is_mu=0
    for i in range(len(ha1)):
        if ha1[i]!=ha2[i] and ha1[i]!="-" and ha2[i]!="-"and ha1[i]!="X" and ha2[i]!="X":
            is_mu=1
            break
    return is_mu

#糖基化位点是否发生突变
def gly_site(ha1,ha2):
#寻找糖基化位点 出现N-X-S/T X不能为P
    is_mut=0
    pos1=find_all(ha1)
    pos2=find_all(ha2)
    pos=list(set(pos1+pos2))
    for i in pos:
        if ha1[i] != ha2[i] or ha1[i+1] != ha2[i+1] or ha1[i+2] != ha2[i+2]:
            is_mut=1
            return is_mut
    return is_mut

#4种氨基酸理化性质
def phy_pro(ha1,ha2,mut_pos):
    phy={'A':[1.8,67,6,41],'R':[-4.5,148,10.76,-14],'N':[-3.5,91,5.41,-28],
         'D':[-3.5,109,2.77,-55],'C':[2.5,86,5.07,49],'Q':[-3.5,90,5.56,-10],
         'E':[-3.5,114,3.22,-31],'G':[-0.4,48,5.97,0],'H':[-3.2,118,7.59,8],
         'I':[4.5,124,6.02,99],'L':[3.8,124,5.98,97],'K':[-3.9,135,9.74,-23],
         'M':[1.9,124,5.74,74],'F':[2.8,135,5.48,100],'P':[-1.6,96,6.3,-46],
         'S':[-0.8,73,5.68,-5],'T':[-0.7,93,5.6,13],'W':[-0.9,163,5.89,97],
         'Y':[-1.3,141,5.66,63],'V':[4.2,105,5.96,76]}
    hyd_list = []
    vol_list = []
    char_list = []
    pol_list = []
    for i in mut_pos:
        h = abs(phy[ha1[i]][0] - phy[ha2[i]][0])
        v = abs(phy[ha1[i]][1] - phy[ha2[i]][1])
        c = abs(phy[ha1[i]][2] - phy[ha2[i]][2])
        p = abs(phy[ha1[i]][3] - phy[ha2[i]][3])
        hyd_list.append(h)
        vol_list.append(v)
        char_list.append(c)
        pol_list.append(p)
    #选变化最大的前三个取均值
    hyd_list.sort(reverse=True)
    vol_list.sort(reverse=True)
    char_list.sort(reverse=True)
    pol_list.sort(reverse=True)
    if len(mut_pos)>3:
        hyd = (hyd_list[0]+hyd_list[1]+hyd_list[2])/3
        vol = (vol_list[0] + vol_list[1] + vol_list[2])/3
        char = (char_list[0] + char_list[1] + char_list[2])/3
        pol = (pol_list[0] + pol_list[1] + pol_list[2])/3
    elif len(mut_pos)>0:
        hyd = sum(hyd_list)/len(mut_pos)
        vol = sum(vol_list)/len(mut_pos)
        char = sum(char_list)/len(mut_pos)
        pol = sum(pol_list)/len(mut_pos)
    else :
        hyd = 0
        vol = 0
        char = 0
        pol = 0

    final_phy=[]
    final_phy.append(hyd)
    final_phy.append(vol)
    final_phy.append(char)
    final_phy.append(pol)
    return final_phy



def find_all(ha):
    start=0
    gly='N.(T|S)'
    pos=[]
    for m in re.finditer(gly, ha):
        if ha[m.start()+1] != "P":
            pos.append(m.start())

    return pos


def cal_feature_2(HAs_1,HAs_2):
    # for i in range(269):
    #     for j in range(i+1,270):
    #         fea=[]
    #         for k in range(328):
    #             if HAs[i][k] == HAs[j][k]:
    #                 fea.append(0)
    #             elif HAs[i][k] != HAs[j][k]:
    #                 fea.append(1)
    #         Feature.append(fea)

    for i in range(300):
        for j in range(2):
            fea=[]
            for k in range(328):
                if HAs_1[i][k] == HAs_2[j][k]:
                    fea.append(0)
                else:
                    fea.append(1)
            Feature.append(fea)


def cal_change_feature():
    table = xlrd.open_workbook("virus_sequence.xls")
    sheet = table.sheet_by_index(0)

    wb = openpyxl.Workbook()
    ws = wb.active

    virus_group={'0':[],'1':[],'2':[],'3':[],'4':[],'5':[],'6':[],'7':[],'8':[],'9':[],'10':[],'11':[]
        , '12': [],'13':[]}
    virus_drift=[7,13,5,0,8,4,2,9,6,3,1]

    for i in range(sheet.nrows):
        print(i)
        ha=sheet.cell(i,2).value
        virus_group[str(int(sheet.cell(i,3).value))].append(ha)


    for i in range(len(virus_drift)-1):
        print(i)
        HA_1 = virus_group.get(str(virus_drift[i]))
        HA_2 = virus_group.get(str(virus_drift[i+1]))
        F=feature_cal(HA_1,HA_2)
        f = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        for i in range(len(F)):
            for j in range(12):
                f[j]=f[j]+F[i][j]
        for i in range(12):
            f[i]=f[i]/len(F)

        ws.append(f)


    # for key in range(14):
    #     print(key)
    #     HA=virus_group.get(str(key))
    #     F=feature_cal(HA)
    #     # print(F)
    #     if len(F)==0:
    #         continue
    #     f=[0,0,0,0,0,0,0,0,0,0,0,0]
    #
    #     for i in range(len(F)):
    #         for j in range(12):
    #             f[j]=f[j]+F[i][j]
    #     for i in range(12):
    #         f[i]=f[i]/len(F)
    #
    #     ws.append(f)

    wb.save("feature_drift.xlsx")


if __name__ =='__main__':
    HAs_1 = []
    HAs_2 = []
    HAs_1,HAs_2=readHA()
    Feature = feature_cal(HAs_1,HAs_2)
    print(Feature)
    workbook = Workbook()
    save_file="Feature.xlsx"
    worksheet = workbook.active
    print(len(Feature))
    for i in range(600):

        l=[]
        for j in range(340):
            l.append(Feature[i][j])
        worksheet.append(l)

    # i=1
    # num=1
    # for row in Feature:
    #     print(i)
    #     # if i%100000==0:
    #     #     workbook.save(filename=save_file)
    #     #     save_file="Feature_1999_2022"+"_"+str(num)+".xls"
    #     #     num=num+1
    #     #     workbook = Workbook()
    #     #     worksheet = workbook.active
    #     #     l = []
    #     #     for i in range(340):
    #     #         l.append(i)
    #     #     worksheet.append(l)
    #     worksheet.append(row)
    #     i=i+1# 把每一行append到worksheet中
    workbook.save(filename=save_file)
    # print(HAs)
    # print(len(HAs))
    # dif=0
    # for i in range(328):
    #     if HAs[1][i]!=HAs[4][i]:
    #         dif=dif+1
    # print(dif)
    # print((328-dif)/328)

    # cal_change_feature()